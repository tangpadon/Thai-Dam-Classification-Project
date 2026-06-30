"""Generate model evaluation summary Excel from Weka output"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="2F5496")
sub_font = Font(bold=True, size=11)
sub_fill = PatternFill("solid", fgColor="D6E4F0")
best_fill = PatternFill("solid", fgColor="C6EFCE")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if fill:
            cell.fill = fill

# ═══════════════════════════════════════════
# Sheet 1 — InfoGain Ranking
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "InfoGain Ranking"

ws1.merge_cells("A1:D1")
ws1.cell(1, 1, "InfoGain Attribute Ranking — 7-Day Forecast").font = Font(bold=True, size=13)
ws1.merge_cells("E1:H1")
ws1.cell(1, 5, "InfoGain Attribute Ranking — 30-Day Forecast").font = Font(bold=True, size=13)

headers = ["Rank", "Attribute", "InfoGain", ""]
for i, h in enumerate(headers, 1):
    ws1.cell(2, i, h)
    ws1.cell(2, i + 4, h)
style_header(ws1, 2, 4)
style_header(ws1, 2, 8)

infogain_7d = [
    (1, "percent_storage", 1.0149),
    (2, "volume", 0.3831),
    (3, "month", 0.3049),
    (4, "id", 0.2953),
    (5, "name", 0.2953),
    (6, "active_storage", 0.2952),
    (7, "capacity", 0.2952),
    (8, "storage", 0.2915),
    (9, "dead_storage", 0.2761),
    (10, "outflow", 0.1907),
    (11, "season", 0.1294),
    (12, "region", 0.0882),
    (13, "inflow", 0.0711),
    (14, "owner", 0.0154),
]

infogain_30d = [
    (1, "percent_storage", 0.638),
    (2, "month", 0.3086),
    (3, "id", 0.2933),
    (4, "name", 0.2933),
    (5, "capacity", 0.2933),
    (6, "active_storage", 0.2932),
    (7, "storage", 0.2896),
    (8, "dead_storage", 0.2734),
    (9, "volume", 0.2705),
    (10, "outflow", 0.2043),
    (11, "inflow", 0.1144),
    (12, "season", 0.0968),
    (13, "region", 0.0898),
    (14, "owner", 0.0155),
]

for i, (r, attr, gain) in enumerate(infogain_7d):
    row = i + 3
    ws1.cell(row, 1, r)
    ws1.cell(row, 2, attr)
    ws1.cell(row, 3, round(gain, 4))
    fill = best_fill if attr in ("percent_storage", "month", "id", "outflow", "inflow") else None
    style_row(ws1, row, 3, fill)

for i, (r, attr, gain) in enumerate(infogain_30d):
    row = i + 3
    ws1.cell(row, 5, r)
    ws1.cell(row, 6, attr)
    ws1.cell(row, 7, round(gain, 4))
    fill = best_fill if attr in ("percent_storage", "month", "id", "outflow", "inflow") else None
    style_row(ws1, row, 7, fill)

# Recommended features note
note_row = len(infogain_7d) + 4
ws1.merge_cells(f"A{note_row}:D{note_row}")
ws1.cell(note_row, 1, "✅ Recommended features: percent_storage + month + id + outflow + inflow").font = Font(bold=True, color="006100")
ws1.merge_cells(f"E{note_row}:H{note_row}")
ws1.cell(note_row, 5, "✅ Recommended features: percent_storage + month + id + outflow + inflow").font = Font(bold=True, color="006100")

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['E'].width = 8
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 12

# ═══════════════════════════════════════════
# Sheet 2 — Model Comparison 7-Day
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Model Comparison 7-Day")

# ── 7-day results ──
models_7d = [
    {
        "name": "J48 (Decision Tree)",
        "correct": 8253, "incorrect": 178, "total": 8431,
        "pct": 97.8887, "kappa": 0.956, "mae": 0.0187, "rmse": 0.113,
        "rae": 5.8729, "rrse": 28.3397,
        "class_detail": [
            ("drought", 0.971, 0.007, 0.937, 0.971, 0.954, 0.949, 0.976, 0.923),
            ("normal",  0.982, 0.027, 0.987, 0.982, 0.984, 0.951, 0.979, 0.983),
            ("flood",   0.974, 0.008, 0.973, 0.974, 0.973, 0.966, 0.990, 0.968),
        ],
        "confusion": [[835,25,0],[56,5649,50],[0,47,1769]],
        "size": "629 leaves, 447 nodes",
        "build_time": "0.11s",
    },
    {
        "name": "RandomForest (100 trees)",
        "correct": 8316, "incorrect": 115, "total": 8431,
        "pct": 98.636, "kappa": 0.9714, "mae": 0.0162, "rmse": 0.0848,
        "raE": 5.1117, "rrse": 21.2624,
        "class_detail": [
            ("drought", 0.970, 0.003, 0.975, 0.970, 0.973, 0.969, 0.998, 0.992),
            ("normal",  0.990, 0.022, 0.990, 0.990, 0.990, 0.969, 0.998, 0.999),
            ("flood",   0.982, 0.005, 0.981, 0.982, 0.981, 0.976, 1.000, 0.998),
        ],
        "confusion": [[834,26,0],[21,5699,35],[0,33,1783]],
        "size": "100 trees",
        "build_time": "1.49s",
    },
    {
        "name": "Logistic Regression",
        "correct": 8119, "incorrect": 312, "total": 8431,
        "pct": 96.2994, "kappa": 0.9223, "mae": 0.0406, "rmse": 0.1387,
        "raE": 12.7711, "rrse": 34.7841,
        "class_detail": [
            ("drought", 0.948, 0.008, 0.932, 0.948, 0.940, 0.933, 0.997, 0.979),
            ("normal",  0.975, 0.062, 0.971, 0.975, 0.973, 0.914, 0.992, 0.996),
            ("flood",   0.933, 0.013, 0.952, 0.933, 0.942, 0.927, 0.996, 0.987),
        ],
        "confusion": [[815,45,0],[59,5610,86],[0,122,1694]],
        "size": "6 features × 3 classes",
        "build_time": "1.75s",
    },
]

# Summary table
row = 1
ws2.cell(row, 1, "Model Performance Summary — 7-Day Forecast (features: percent_storage, inflow, outflow, month, id)")
ws2.cell(row, 1).font = Font(bold=True, size=13)
ws2.merge_cells(f"A{row}:K{row}")
row = 2
sum_headers = ["Model", "Correct", "Incorrect", "Accuracy %", "Kappa", "MAE", "RMSE", "RAE %", "RRSE %", "Tree Size", "Build Time"]
for c, h in enumerate(sum_headers, 1):
    ws2.cell(row, c, h)
style_header(ws2, row, len(sum_headers))

for i, m in enumerate(models_7d):
    r = row + 1 + i
    vals = [m["name"], m["correct"], m["incorrect"], m["pct"], m["kappa"],
            m["mae"], m["rmse"], m.get("raE", m.get("rae")), m["rrse"],
            m["size"], m["build_time"]]
    for c, v in enumerate(vals, 1):
        ws2.cell(r, c, v)
    fill = best_fill if i == 1 else None  # best = RandomForest
    style_row(ws2, r, len(sum_headers), fill)

# Class detail
row = row + len(models_7d) + 2
ws2.cell(row, 1, "Per-Class Detailed Accuracy").font = Font(bold=True, size=12)
ws2.merge_cells(f"A{row}:I{row}")
row += 1
cls_headers = ["Model", "Class", "TP Rate", "FP Rate", "Precision", "Recall", "F-Measure", "MCC", "ROC Area"]
for c, h in enumerate(cls_headers, 1):
    ws2.cell(row, c, h)
style_header(ws2, row, len(cls_headers))

for m in models_7d:
    for cls_data in m["class_detail"]:
        r = row + 1
        ws2.cell(r, 1, m["name"])
        for c, v in enumerate(cls_data, 2):
            ws2.cell(r, c, v)
        style_row(ws2, r, len(cls_headers))
        # Mark best ROC per class
        row += 1

# Confusion matrices
row += 1
ws2.cell(row, 1, "Confusion Matrices (actual rows × predicted cols: drought, normal, flood)").font = Font(bold=True, size=12)
ws2.merge_cells(f"A{row}:D{row}")
row += 1

for m in models_7d:
    ws2.cell(row, 1, m["name"]).font = sub_font
    ws2.cell(row, 1).fill = sub_fill
    ws2.merge_cells(f"A{row}:D{row}")
    row += 1
    cm = m["confusion"]
    for cm_row in cm:
        ws2.cell(row, 1, cm_row[0])  # actual label shown as count
        ws2.cell(row, 2, cm_row[1])
        ws2.cell(row, 3, cm_row[2])
        style_row(ws2, row, 3)
        row += 1
    row += 1

ws2.column_dimensions['A'].width = 25

# ═══════════════════════════════════════════
# Sheet 3 — Model Comparison 30-Day
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Model Comparison 30-Day")

models_30d = [
    {
        "name": "J48 (Decision Tree)",
        "correct": 8235, "incorrect": 196, "total": 8431,
        "pct": 97.6752, "kappa": 0.9516, "mae": 0.0193, "rmse": 0.1195,
        "rae": 6.0019, "rrse": 29.7732,
        "class_detail": [
            ("drought", 0.945, 0.004, 0.962, 0.945, 0.953, 0.948, 0.981, 0.942),
            ("normal",  0.987, 0.044, 0.979, 0.987, 0.983, 0.947, 0.979, 0.983),
            ("flood",   0.961, 0.007, 0.976, 0.961, 0.968, 0.959, 0.986, 0.958),
        ],
        "confusion": [[807,46,1],[32,5622,43],[0,74,1806]],
        "size": "1061 leaves, 729 nodes",
        "build_time": "0.09s",
    },
    {
        "name": "RandomForest (100 trees)",
        "correct": 8308, "incorrect": 123, "total": 8431,
        "pct": 98.5411, "kappa": 0.9698, "mae": 0.0216, "rmse": 0.0922,
        "rae": 6.7395, "rrse": 22.9548,
        "class_detail": [
            ("drought", 0.970, 0.002, 0.980, 0.970, 0.975, 0.972, 0.999, 0.995),
            ("normal",  0.990, 0.025, 0.988, 0.990, 0.989, 0.967, 0.998, 0.999),
            ("flood",   0.978, 0.006, 0.979, 0.978, 0.979, 0.973, 0.999, 0.997),
        ],
        "confusion": [[828,26,0],[17,5641,39],[0,41,1839]],
        "size": "100 trees",
        "build_time": "1.53s",
    },
    {
        "name": "Logistic Regression",
        "correct": 7356, "incorrect": 1075, "total": 8431,
        "pct": 87.2494, "kappa": 0.7304, "mae": 0.1236, "rmse": 0.2457,
        "rae": 38.4738, "rrse": 61.2022,
        "class_detail": [
            ("drought", 0.738, 0.023, 0.782, 0.738, 0.759, 0.733, 0.979, 0.859),
            ("normal",  0.922, 0.230, 0.893, 0.922, 0.907, 0.705, 0.940, 0.971),
            ("flood",   0.785, 0.041, 0.845, 0.785, 0.814, 0.764, 0.967, 0.882),
        ],
        "confusion": [[630,224,0],[176,5251,270],[0,405,1475]],
        "size": "6 features × 3 classes",
        "build_time": "1.35s",
    },
]

# Summary table
row = 1
ws3.cell(row, 1, "Model Performance Summary — 30-Day Forecast (features: percent_storage, inflow, outflow, month, id)")
ws3.cell(row, 1).font = Font(bold=True, size=13)
ws3.merge_cells(f"A{row}:K{row}")
row = 2
for c, h in enumerate(sum_headers, 1):
    ws3.cell(row, c, h)
style_header(ws3, row, len(sum_headers))

for i, m in enumerate(models_30d):
    r = row + 1 + i
    vals = [m["name"], m["correct"], m["incorrect"], m["pct"], m["kappa"],
            m["mae"], m["rmse"], m["rae"], m["rrse"],
            m["size"], m["build_time"]]
    for c, v in enumerate(vals, 1):
        ws3.cell(r, c, v)
    fill = best_fill if i == 1 else None
    style_row(ws3, r, len(sum_headers), fill)

# Class detail
row = row + len(models_30d) + 2
ws3.cell(row, 1, "Per-Class Detailed Accuracy").font = Font(bold=True, size=12)
ws3.merge_cells(f"A{row}:I{row}")
row += 1
for c, h in enumerate(cls_headers, 1):
    ws3.cell(row, c, h)
style_header(ws3, row, len(cls_headers))

for m in models_30d:
    for cls_data in m["class_detail"]:
        r = row + 1
        ws3.cell(r, 1, m["name"])
        for c, v in enumerate(cls_data, 2):
            ws3.cell(r, c, v)
        style_row(ws3, r, len(cls_headers))
        row += 1

# Confusion matrices
row += 1
ws3.cell(row, 1, "Confusion Matrices (actual rows × predicted cols: drought, normal, flood)").font = Font(bold=True, size=12)
ws3.merge_cells(f"A{row}:D{row}")
row += 1

for m in models_30d:
    ws3.cell(row, 1, m["name"]).font = sub_font
    ws3.cell(row, 1).fill = sub_fill
    ws3.merge_cells(f"A{row}:D{row}")
    row += 1
    cm = m["confusion"]
    for cm_row in cm:
        ws3.cell(row, 1, cm_row[0])
        ws3.cell(row, 2, cm_row[1])
        ws3.cell(row, 3, cm_row[2])
        style_row(ws3, row, 3)
        row += 1
    row += 1

ws3.column_dimensions['A'].width = 25

# ═══════════════════════════════════════════
# Sheet 4 — Feature Set Comparison
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Feature Set Comparison")

ws4.cell(1, 1, "Feature Set Comparison — Before vs After InfoGain").font = Font(bold=True, size=13)
ws4.merge_cells("A1:F1")

headers_fs = ["", "Original (14 features)", "", "Selected (5 features)", "", ""]
row = 2
for c, h in enumerate(headers_fs, 1):
    ws4.cell(row, c, h)
style_header(ws4, row, 6)

ws4.cell(row+1, 1, "Attribute")
ws4.cell(row+1, 2, "InfoGain 7d")
ws4.cell(row+1, 3, "InfoGain 30d")
ws4.cell(row+1, 4, "Attribute")
ws4.cell(row+1, 5, "InfoGain 7d")
ws4.cell(row+1, 6, "InfoGain 30d")
style_header(ws4, row+1, 6)

all_attrs = [
    ("capacity", 0.2952, 0.2933, False),
    ("storage", 0.2915, 0.2896, False),
    ("active_storage", 0.2952, 0.2932, False),
    ("dead_storage", 0.2761, 0.2734, False),
    ("volume", 0.3831, 0.2705, False),
    ("percent_storage", 1.0149, 0.638, True),
    ("inflow", 0.0711, 0.1144, True),
    ("outflow", 0.1907, 0.2043, True),
    ("month", 0.3049, 0.3086, True),
    ("id", 0.2953, 0.2933, True),
    ("name", 0.2953, 0.2933, False),
    ("region", 0.0882, 0.0898, False),
    ("owner", 0.0154, 0.0155, False),
    ("season", 0.1294, 0.0968, False),
]

selected = [a for a in all_attrs if a[3]]
dropped = [a for a in all_attrs if not a[3]]

for i, (attr, g7, g30, _) in enumerate(dropped):
    r = row + 2 + i
    ws4.cell(r, 1, attr)
    ws4.cell(r, 2, round(g7, 4))
    ws4.cell(r, 3, round(g30, 4))
    style_row(ws4, r, 3)

for i, (attr, g7, g30, _) in enumerate(selected):
    r = row + 2 + i
    ws4.cell(r, 4, attr)
    ws4.cell(r, 5, round(g7, 4))
    ws4.cell(r, 6, round(g30, 4))
    fill = best_fill
    style_row(ws4, r, 6, fill)

# Redundant groups
r = row + 2 + max(len(dropped), len(selected)) + 2
ws4.cell(r, 1, "Redundancy Groups Detected").font = Font(bold=True, size=12, color="C00000")
ws4.merge_cells(f"A{r}:F{r}")
r += 1
groups = [
    ("Group 1 (เขื่อน identity)", "id, name, capacity, storage, active_storage, dead_storage, volume", "ใช้แค่ id ตัวเดียว"),
    ("Group 2 (ฤดูกาล)", "month, season", "month ให้ granularity ดีกว่า"),
    ("Group 3 (ภูมิภาค)", "region → ซ้ำกับ id", "id บอกอยู่แล้วว่าเขื่อนไหนอยู่ภาคไหน"),
    ("Group 4 (เจ้าของ)", "owner → ทุกเขื่อนเป็นกรมชลประทาน", "แทบไม่มีข้อมูล"),
]
for g_name, g_members, g_note in groups:
    ws4.cell(r, 1, g_name).font = Font(bold=True)
    ws4.cell(r, 2, g_members)
    ws4.cell(r, 4, g_note)
    ws4.cell(r, 4).font = Font(color="006100")
    ws4.merge_cells(f"B{r}:C{r}")
    ws4.merge_cells(f"D{r}:F{r}")
    style_row(ws4, r, 6)
    r += 1

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 14
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 14
ws4.column_dimensions['F'].width = 14

# ═══════════════════════════════════════════
# Sheet 5 — Recommendation
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("Recommendation")
ws5.cell(1, 1, "Model Recommendation").font = Font(bold=True, size=14, color="2F5496")

recs = [
    ("🏆 Best Overall (7d & 30d)", "RandomForest (100 trees)", "98.6% / 98.5% accuracy, ROC ≈ 0.999, robust, ไม่ overfit"),
    ("🥈 Fast & Interpretable", "J48 Decision Tree", "97.9% / 97.7% accuracy, build 0.1s, อ่านกฏ决策 tree ได้"),
    ("❌ Not Recommended", "Logistic Regression", "87.2% (30d) drop ชัด, linear assumption ไม่เหมาะกับ data"),
    ("", "", ""),
    ("🎯 Recommended Feature Set (5 features)", "", "percent_storage + month + id + outflow + inflow"),
    ("📉 Dropped (redundant/low gain)", "", "capacity, storage, active_storage, dead_storage, volume, name, region, owner, season"),
    ("", "", ""),
    ("💡 Next Steps", "", ""),
    ("1.", "Retrain final model with 5 features", "ลด complexity, ลด encoding problems, เพิ่ม generalization"),
    ("2.", "Use RandomForest for production", "ROC ≈ 0.999, class balance ดี, ทน missing value"),
    ("3.", "Deploy model via PMML or serialized .model", "save build time ที่ 1.5s ไม่ต้อง train ซ้ำตอน deploy"),
]

for i, (col1, col2, col3) in enumerate(recs):
    r = i + 3
    ws5.cell(r, 1, col1)
    ws5.cell(r, 2, col2)
    ws5.cell(r, 3, col3)
    if col1.startswith(("🏆", "🥈", "❌", "🎯", "📉", "💡")):
        ws5.cell(r, 1).font = Font(bold=True, color="2F5496", size=11)
    style_row(ws5, r, 3)

ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 35
ws5.column_dimensions['C'].width = 55

# ── Save ──
path = "C:\\dam-forecast-project\\model_evaluation_summary.xlsx"
wb.save(path)
print(f"✅ Saved to {path}")
